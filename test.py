import openpyxl

wb = openpyxl.load_workbook('D:\download\实施工程师名单.xlsx')

sheet = wb.get_sheet_by_name("Sheet1")
# print(sheet.max_row)
# row = sheet.max_row
# print(sheet.max_column)
# column = sheet.max_column
print('replace into mediation_tjs_user_auth (oa_id, name, user) values')
for i in sheet["D"]:
    if i.value is not None:
        user = i.value
        # print('((select customer.customer_id from customer where customer.username="' + user + '"), (select customer.name from customer where customer.username="' + user + '"), "' + user + '"),')

        print('"'+ i.value + '", ')



s = 'http://117.90.137.91:9000'

print(s.split(':')[1].split('//')[1])
